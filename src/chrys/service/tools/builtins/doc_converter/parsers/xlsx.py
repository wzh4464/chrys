# Copyright (c) 2026 Chrys. All rights reserved.

"""XLSX parser — converts Excel spreadsheets to Markdown using openpyxl."""

from __future__ import annotations

from chrys.service.tools.builtins.doc_converter.parsers._table import rows_to_markdown_table
from chrys.service.tools.builtins.doc_converter.parsers.base import DocumentImageSink, ParsedDocument


class XlsxParser:
    """Convert XLSX files to Markdown tables via ``openpyxl``."""

    @property
    def supported_extensions(self) -> frozenset[str]:
        return frozenset({".xlsx"})

    def parse(self, path: str, *, image_sink: DocumentImageSink | None = None) -> ParsedDocument:
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        parts: list[str] = []
        try:
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                parts.append(f"# Sheet: {sheet_name}")
                if not rows:
                    parts.append("(empty sheet)")
                    continue
                parts.append(rows_to_markdown_table(rows))
        finally:
            wb.close()
        return ParsedDocument(markdown="\n\n".join(parts))
